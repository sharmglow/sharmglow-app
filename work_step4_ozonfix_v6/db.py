"""
db.py — SQLite database layer for SmartSklad
Файл базы данных: smartsklad.db в LocalAppData/SmartSklad/data
"""
import sqlite3, os, json, shutil
from pathlib import Path
from datetime import date

APP_NAME = "SmartSklad"
DB_FILENAME = "smartsklad.db"
LEGACY_DB_FILENAME = "sharmglow.db"


def _default_data_dir() -> Path:
    local_appdata = os.getenv("LOCALAPPDATA")
    if local_appdata:
        return Path(local_appdata) / APP_NAME / "data"
    return Path.home() / ".smartsklad" / "data"


def _default_db_path() -> Path:
    return _default_data_dir() / DB_FILENAME


DB_PATH = _default_db_path()
LEGACY_DB_PATH = Path(os.path.dirname(os.path.abspath(__file__))) / LEGACY_DB_FILENAME


class DB:
    def __init__(self, path=None):
        self.path = Path(path) if path else DB_PATH
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._migrate_legacy_db_if_needed()
        self.conn = sqlite3.connect(str(self.path))
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA foreign_keys=ON")
        self._migrate()

    def _migrate_legacy_db_if_needed(self):
        if self.path.exists():
            return
        if self.path == LEGACY_DB_PATH:
            return
        if not LEGACY_DB_PATH.exists():
            return

        # Пытаемся аккуратно перенести старую БД из папки проекта в постоянную папку AppData.
        try:
            src = sqlite3.connect(str(LEGACY_DB_PATH))
            try:
                src.execute("PRAGMA wal_checkpoint(FULL)")
            except Exception:
                pass
            dst = sqlite3.connect(str(self.path))
            try:
                src.backup(dst)
                dst.commit()
            finally:
                dst.close()
                src.close()
        except Exception:
            # Фолбэк: обычное копирование файлов, если backup по какой-то причине недоступен.
            shutil.copy2(LEGACY_DB_PATH, self.path)
            for suffix in ("-wal", "-shm"):
                legacy_side = LEGACY_DB_PATH.with_name(LEGACY_DB_PATH.name + suffix)
                target_side = self.path.with_name(self.path.name + suffix)
                if legacy_side.exists():
                    shutil.copy2(legacy_side, target_side)

    def _migrate(self):
        c = self.conn
        c.executescript("""
        CREATE TABLE IF NOT EXISTS warehouses (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            type TEXT DEFAULT 'Собственный',
            note TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS products (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            art       TEXT UNIQUE NOT NULL,
            name      TEXT NOT NULL,
            cat       TEXT DEFAULT '',
            unit      TEXT DEFAULT 'шт',
            min_stock INTEGER DEFAULT 0,
            bc_wb     TEXT DEFAULT '',
            bc_ozon   TEXT DEFAULT '',
            supplier  TEXT DEFAULT '',
            note      TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS arrivals (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            dt         TEXT NOT NULL,
            art        TEXT NOT NULL,
            wh_name    TEXT NOT NULL,
            qty        INTEGER NOT NULL,
            supplier   TEXT DEFAULT '',
            note       TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS moves (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            dt         TEXT NOT NULL,
            art        TEXT NOT NULL,
            from_wh    TEXT NOT NULL,
            to_wh      TEXT NOT NULL,
            qty        INTEGER NOT NULL,
            barcode    TEXT DEFAULT '',
            mp_type    TEXT DEFAULT '',
            note       TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS label_templates (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            name      TEXT NOT NULL,
            is_default INTEGER DEFAULT 0,
            config    TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS marketplace_accounts (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            marketplace_code TEXT NOT NULL,
            account_name     TEXT NOT NULL,
            api_key          TEXT DEFAULT '',
            client_id        TEXT DEFAULT '',
            extra_json       TEXT DEFAULT '{}',
            is_active        INTEGER DEFAULT 1,
            created_at       TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at       TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(marketplace_code, account_name)
        );

        CREATE TABLE IF NOT EXISTS product_barcodes (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id       INTEGER NOT NULL,
            barcode          TEXT NOT NULL,
            source           TEXT DEFAULT '',
            marketplace_code TEXT DEFAULT '',
            is_primary       INTEGER DEFAULT 0,
            created_at       TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(product_id, barcode),
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS marketplace_product_links (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id          INTEGER NOT NULL,
            marketplace_code    TEXT NOT NULL,
            account_id          INTEGER,
            external_product_id TEXT DEFAULT '',
            external_sku        TEXT DEFAULT '',
            external_offer_id   TEXT DEFAULT '',
            vendor_code         TEXT DEFAULT '',
            external_name       TEXT DEFAULT '',
            raw_payload_json    TEXT DEFAULT '{}',
            synced_at           TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
            FOREIGN KEY(account_id) REFERENCES marketplace_accounts(id) ON DELETE SET NULL
        );

        CREATE INDEX IF NOT EXISTS idx_product_barcodes_barcode ON product_barcodes(barcode);
        CREATE INDEX IF NOT EXISTS idx_marketplace_accounts_code ON marketplace_accounts(marketplace_code);
        CREATE INDEX IF NOT EXISTS idx_marketplace_links_product ON marketplace_product_links(product_id);
        CREATE INDEX IF NOT EXISTS idx_marketplace_links_code ON marketplace_product_links(marketplace_code);
        """)

        # Seed default warehouses if empty
        if not c.execute("SELECT 1 FROM warehouses LIMIT 1").fetchone():
            c.executemany("INSERT INTO warehouses(code,name,type,note) VALUES(?,?,?,?)", [
                ("ТВ-2",  "Склад Тверская 2 (основной)", "Центральный",      "Приход из Китая"),
                ("WB-1",  "Склад Wildberries",           "Маркетплейс WB",   ""),
                ("OZ-1",  "Склад OZON",                  "Маркетплейс OZON", ""),
                ("ТВ-3",  "Склад Тверская 3",            "Собственный",      ""),
            ])

        # Seed default label template if empty
        if not c.execute("SELECT 1 FROM label_templates LIMIT 1").fetchone():
            default_cfg = json.dumps({
                "width_mm": 58, "height_mm": 40,
                "bg_color": "#ffffff", "border": True,
                "blocks": [
                    {"type": "mp_badge",  "x": 0.72, "y": 0.04, "font_size": 7,  "bold": True,  "visible": True},
                    {"type": "name",      "x": 0.05, "y": 0.05, "font_size": 9,  "bold": True,  "visible": True, "align": "center", "max_lines": 2},
                    {"type": "divider",   "x": 0.05, "y": 0.35, "visible": True, "color": "#cccccc"},
                    {"type": "art",       "x": 0.05, "y": 0.40, "font_size": 11, "bold": True,  "visible": True, "italic": True, "font": "serif"},
                    {"type": "logo",      "x": 0.60, "y": 0.40, "font_size": 9,  "bold": False, "visible": True, "italic": True, "font": "serif"},
                    {"type": "barcode",   "x": 0.05, "y": 0.58, "visible": True, "height_ratio": 0.30},
                    {"type": "bc_number", "x": 0.05, "y": 0.93, "font_size": 6,  "bold": False, "visible": True, "align": "center"},
                    {"type": "price",     "x": 0.05, "y": 0.82, "font_size": 9,  "bold": True,  "visible": False, "align": "right"},
                ]
            })
            c.execute("INSERT INTO label_templates(name,is_default,config) VALUES(?,1,?)",
                      ("Стандарт 58×40", default_cfg))

        self._backfill_marketplace_foundation()
        c.commit()

    def _backfill_marketplace_foundation(self):
        rows = self.conn.execute(
            "SELECT id, art, name, bc_wb, bc_ozon FROM products ORDER BY id"
        ).fetchall()
        for row in rows:
            product_id = row["id"]
            art = (row["art"] or "").strip()
            name = (row["name"] or "").strip()
            bc_wb = (row["bc_wb"] or "").strip()
            bc_ozon = (row["bc_ozon"] or "").strip()

            if bc_wb:
                self.conn.execute(
                    "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
                    (product_id, bc_wb, "legacy_products", "wb", 1),
                )
                self.conn.execute(
                    "INSERT OR IGNORE INTO marketplace_product_links(product_id, marketplace_code, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?)",
                    (product_id, "wb", bc_wb, art, name, json.dumps({"seed": "legacy_bc_wb"}, ensure_ascii=False)),
                )

            if bc_ozon:
                self.conn.execute(
                    "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
                    (product_id, bc_ozon, "legacy_products", "ozon", 1),
                )
                self.conn.execute(
                    "INSERT OR IGNORE INTO marketplace_product_links(product_id, marketplace_code, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?)",
                    (product_id, "ozon", bc_ozon, art, name, json.dumps({"seed": "legacy_bc_ozon"}, ensure_ascii=False)),
                )

    def close(self):
        self.conn.close()

    # ── WAREHOUSES ──────────────────────────────────────────
    def get_warehouses(self):
        return self.conn.execute("SELECT * FROM warehouses ORDER BY id").fetchall()

    def add_warehouse(self, code, name, wtype, note=""):
        self.conn.execute("INSERT INTO warehouses(code,name,type,note) VALUES(?,?,?,?)",
                          (code, name, wtype, note))
        self.conn.commit()

    def update_warehouse(self, wid, code, name, wtype, note):
        self.conn.execute("UPDATE warehouses SET code=?,name=?,type=?,note=? WHERE id=?",
                          (code, name, wtype, note, wid))
        self.conn.commit()

    def delete_warehouse(self, wid):
        self.conn.execute("DELETE FROM warehouses WHERE id=?", (wid,))
        self.conn.commit()

    def mp_type_for_wh(self, wh_name):
        n = wh_name.upper()
        if "WB" in n or "WILDBERRIES" in n: return "WB"
        if "OZON" in n or "ОЗОН" in n:      return "OZON"
        return ""

    # ── PRODUCTS ────────────────────────────────────────────
    def get_products(self, search=""):
        q = f"%{search}%"
        return self.conn.execute(
            "SELECT * FROM products WHERE art LIKE ? OR name LIKE ? OR cat LIKE ? ORDER BY art",
            (q, q, q)).fetchall()

    def get_product(self, art):
        return self.conn.execute("SELECT * FROM products WHERE art=?", (art,)).fetchone()

    def get_product_by_barcode(self, bc):
        row = self.conn.execute(
            "SELECT p.* FROM product_barcodes b "
            "JOIN products p ON p.id=b.product_id WHERE b.barcode=? LIMIT 1",
            (bc,),
        ).fetchone()
        if row:
            return row
        return self.conn.execute(
            "SELECT * FROM products WHERE bc_wb=? OR bc_ozon=?", (bc, bc)).fetchone()

    def add_product(self, art, name, cat, unit, min_stock, bc_wb, bc_ozon, supplier, note):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO products(art,name,cat,unit,min_stock,bc_wb,bc_ozon,supplier,note) VALUES(?,?,?,?,?,?,?,?,?)",
            (art, name, cat, unit, min_stock, bc_wb, bc_ozon, supplier, note))
        pid = cur.lastrowid
        self.conn.commit()
        self._sync_legacy_product_marketplace_data(pid, art, name, bc_wb, bc_ozon)
        self.conn.commit()
        return pid

    def update_product(self, pid, art, name, cat, unit, min_stock, bc_wb, bc_ozon, supplier, note):
        self.conn.execute(
            "UPDATE products SET art=?,name=?,cat=?,unit=?,min_stock=?,bc_wb=?,bc_ozon=?,supplier=?,note=? WHERE id=?",
            (art, name, cat, unit, min_stock, bc_wb, bc_ozon, supplier, note, pid))
        self.conn.commit()
        self._sync_legacy_product_marketplace_data(pid, art, name, bc_wb, bc_ozon)
        self.conn.commit()

    def _sync_legacy_product_marketplace_data(self, pid, art, name, bc_wb, bc_ozon):
        # Удаляем старые legacy-значения и заново переносим текущие поля bc_wb / bc_ozon
        self.conn.execute(
            "DELETE FROM product_barcodes WHERE product_id=? AND source='legacy_products' AND marketplace_code IN ('wb','ozon')",
            (pid,),
        )
        self.conn.execute(
            "DELETE FROM marketplace_product_links WHERE product_id=? AND raw_payload_json LIKE ?",
            (pid, '%legacy_bc_%'),
        )
        if bc_wb:
            self.conn.execute(
                "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
                (pid, bc_wb, "legacy_products", "wb", 1),
            )
            self.conn.execute(
                "INSERT OR IGNORE INTO marketplace_product_links(product_id, marketplace_code, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?)",
                (pid, "wb", bc_wb, art, name, json.dumps({"seed": "legacy_bc_wb"}, ensure_ascii=False)),
            )
        if bc_ozon:
            self.conn.execute(
                "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
                (pid, bc_ozon, "legacy_products", "ozon", 1),
            )
            self.conn.execute(
                "INSERT OR IGNORE INTO marketplace_product_links(product_id, marketplace_code, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?)",
                (pid, "ozon", bc_ozon, art, name, json.dumps({"seed": "legacy_bc_ozon"}, ensure_ascii=False)),
            )

    def delete_product(self, pid):
        self.conn.execute("DELETE FROM products WHERE id=?", (pid,))
        self.conn.commit()

    def get_categories(self):
        rows = self.conn.execute("SELECT DISTINCT cat FROM products WHERE cat!='' ORDER BY cat").fetchall()
        return [r[0] for r in rows]

    # ── MARKETPLACE FOUNDATION ─────────────────────────────
    def get_marketplace_accounts(self, marketplace_code=None, only_active=False):
        sql = "SELECT * FROM marketplace_accounts WHERE 1=1"
        params = []
        if marketplace_code:
            sql += " AND marketplace_code=?"
            params.append(marketplace_code)
        if only_active:
            sql += " AND is_active=1"
        sql += " ORDER BY marketplace_code, account_name"
        return self.conn.execute(sql, params).fetchall()

    def save_marketplace_account(self, account_id, marketplace_code, account_name, api_key="", client_id="", extra_json="{}", is_active=1):
        if account_id:
            self.conn.execute(
                "UPDATE marketplace_accounts SET marketplace_code=?, account_name=?, api_key=?, client_id=?, extra_json=?, is_active=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (marketplace_code, account_name, api_key, client_id, extra_json or "{}", int(bool(is_active)), account_id),
            )
            self.conn.commit()
            return account_id

        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO marketplace_accounts(marketplace_code, account_name, api_key, client_id, extra_json, is_active) VALUES(?,?,?,?,?,?)",
            (marketplace_code, account_name, api_key, client_id, extra_json or "{}", int(bool(is_active))),
        )
        self.conn.commit()
        return cur.lastrowid

    def delete_marketplace_account(self, account_id):
        self.conn.execute("DELETE FROM marketplace_accounts WHERE id=?", (account_id,))
        self.conn.commit()

    def get_product_barcodes(self, product_id):
        return self.conn.execute(
            "SELECT * FROM product_barcodes WHERE product_id=? ORDER BY is_primary DESC, id ASC",
            (product_id,),
        ).fetchall()

    def add_product_barcode(self, product_id, barcode, source="", marketplace_code="", is_primary=0):
        self.conn.execute(
            "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
            (product_id, str(barcode).strip(), source, marketplace_code, int(bool(is_primary))),
        )
        self.conn.commit()

    def replace_product_barcodes(self, product_id, barcodes, source="", marketplace_code=""):
        self.conn.execute(
            "DELETE FROM product_barcodes WHERE product_id=? AND source=? AND marketplace_code=?",
            (product_id, source, marketplace_code),
        )
        cleaned = []
        for b in barcodes or []:
            b = str(b).strip()
            if b and b not in cleaned:
                cleaned.append(b)
        for idx, b in enumerate(cleaned):
            self.conn.execute(
                "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,?)",
                (product_id, b, source, marketplace_code, 1 if idx == 0 else 0),
            )
        self.conn.commit()

    def get_marketplace_product_links(self, product_id=None, marketplace_code=None):
        sql = (
            "SELECT l.*, p.art, p.name, a.account_name "
            "FROM marketplace_product_links l "
            "JOIN products p ON p.id=l.product_id "
            "LEFT JOIN marketplace_accounts a ON a.id=l.account_id WHERE 1=1"
        )
        params = []
        if product_id is not None:
            sql += " AND l.product_id=?"
            params.append(product_id)
        if marketplace_code:
            sql += " AND l.marketplace_code=?"
            params.append(marketplace_code)
        sql += " ORDER BY l.marketplace_code, l.id DESC"
        return self.conn.execute(sql, params).fetchall()


    def delete_marketplace_product_link(self, link_id):
        self.conn.execute("DELETE FROM marketplace_product_links WHERE id=?", (link_id,))
        self.conn.commit()

    def delete_duplicate_marketplace_links(self, product_id, marketplace_code, external_product_id="", external_sku="", external_offer_id=""):
        rows = self.conn.execute(
            "SELECT id FROM marketplace_product_links WHERE product_id=? AND marketplace_code=? AND COALESCE(external_product_id,'')=? AND COALESCE(external_sku,'')=? AND COALESCE(external_offer_id,'')=? ORDER BY id ASC",
            (product_id, marketplace_code, external_product_id or "", external_sku or "", external_offer_id or ""),
        ).fetchall()
        if len(rows) <= 1:
            return 0
        keep_id = rows[0]["id"]
        delete_ids = [row["id"] for row in rows[1:]]
        self.conn.executemany("DELETE FROM marketplace_product_links WHERE id=?", [(i,) for i in delete_ids])
        self.conn.commit()
        return len(delete_ids)

    def upsert_marketplace_product_link(
        self,
        product_id,
        marketplace_code,
        account_id=None,
        external_product_id="",
        external_sku="",
        external_offer_id="",
        vendor_code="",
        external_name="",
        raw_payload_json="{}",
    ):
        ext_pid = external_product_id or ""
        ext_sku = external_sku or ""
        ext_offer = external_offer_id or ""

        # 1) Строгое совпадение по полному ключу.
        row = self.conn.execute(
            "SELECT id FROM marketplace_product_links WHERE product_id=? AND marketplace_code=? AND COALESCE(external_offer_id,'')=? AND COALESCE(external_sku,'')=? AND COALESCE(external_product_id,'')=?",
            (product_id, marketplace_code, ext_offer, ext_sku, ext_pid),
        ).fetchone()

        # 2) Более мягкая дедупликация: если для того же товара и маркетплейса уже есть связь
        # с таким же external_product_id / offer_id / sku, обновляем ее, а не создаем новую.
        if row is None:
            clauses = []
            params = [product_id, marketplace_code]
            if ext_pid:
                clauses.append("COALESCE(external_product_id,'')=?")
                params.append(ext_pid)
            if ext_offer:
                clauses.append("COALESCE(external_offer_id,'')=?")
                params.append(ext_offer)
            if ext_sku:
                clauses.append("COALESCE(external_sku,'')=?")
                params.append(ext_sku)
            if clauses:
                sql = (
                    "SELECT id FROM marketplace_product_links WHERE product_id=? AND marketplace_code=? AND ("
                    + " OR ".join(clauses) + ") ORDER BY id ASC LIMIT 1"
                )
                row = self.conn.execute(sql, params).fetchone()

        if row:
            self.conn.execute(
                "UPDATE marketplace_product_links SET account_id=?, external_product_id=?, external_sku=?, external_offer_id=?, vendor_code=?, external_name=?, raw_payload_json=?, synced_at=CURRENT_TIMESTAMP WHERE id=?",
                (account_id, ext_pid, ext_sku, ext_offer, vendor_code, external_name, raw_payload_json or "{}", row["id"]),
            )
            self.conn.commit()
            self.delete_duplicate_marketplace_links(product_id, marketplace_code, ext_pid, ext_sku, ext_offer)
            return row["id"]

        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO marketplace_product_links(product_id, marketplace_code, account_id, external_product_id, external_sku, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?,?,?,?)",
            (
                product_id,
                marketplace_code,
                account_id,
                ext_pid,
                ext_sku,
                ext_offer,
                vendor_code or "",
                external_name or "",
                raw_payload_json or "{}",
            ),
        )
        self.conn.commit()
        self.delete_duplicate_marketplace_links(product_id, marketplace_code, ext_pid, ext_sku, ext_offer)
        return cur.lastrowid



    def get_product_marketplace_map(self, product_id):
        rows = self.conn.execute(
            "SELECT marketplace_code, barcode, source, is_primary FROM product_barcodes WHERE product_id=? ORDER BY marketplace_code, is_primary DESC, id ASC",
            (product_id,),
        ).fetchall()
        result = {}
        for row in rows:
            code = (row["marketplace_code"] or "").strip().lower()
            if not code or code in result:
                continue
            result[code] = row["barcode"] or ""
        return result

    def get_product_marketplace_bindings(self, product_id):
        rows = self.conn.execute(
            "SELECT marketplace_code, barcode, source, is_primary FROM product_barcodes WHERE product_id=? ORDER BY marketplace_code, is_primary DESC, id ASC",
            (product_id,),
        ).fetchall()
        link_rows = self.conn.execute(
            "SELECT marketplace_code, external_sku, external_offer_id, vendor_code, external_name FROM marketplace_product_links WHERE product_id=? ORDER BY id DESC",
            (product_id,),
        ).fetchall()
        data = {}
        for row in rows:
            code = (row["marketplace_code"] or "").strip().lower()
            if not code:
                continue
            entry = data.setdefault(code, {"barcode": "", "external_sku": "", "external_offer_id": "", "vendor_code": "", "external_name": ""})
            if not entry["barcode"]:
                entry["barcode"] = row["barcode"] or ""
        for row in link_rows:
            code = (row["marketplace_code"] or "").strip().lower()
            if not code:
                continue
            entry = data.setdefault(code, {"barcode": "", "external_sku": "", "external_offer_id": "", "vendor_code": "", "external_name": ""})
            if not entry["external_sku"]:
                entry["external_sku"] = row["external_sku"] or ""
            if not entry["external_offer_id"]:
                entry["external_offer_id"] = row["external_offer_id"] or ""
            if not entry["vendor_code"]:
                entry["vendor_code"] = row["vendor_code"] or ""
            if not entry["external_name"]:
                entry["external_name"] = row["external_name"] or ""
        return data

    def save_product_marketplace_bindings(self, product_id, bindings):
        product = self.conn.execute("SELECT art, name FROM products WHERE id=?", (product_id,)).fetchone()
        if not product:
            raise ValueError("Товар не найден")

        allowed = {"wb", "ozon", "yandex_market", "aliexpress"}
        normalized = {}
        for item in bindings or []:
            code = str(item.get("marketplace_code", "")).strip().lower()
            if code not in allowed:
                continue
            normalized[code] = {
                "barcode": str(item.get("barcode", "") or "").strip(),
                "external_sku": str(item.get("external_sku", "") or "").strip(),
                "external_offer_id": str(item.get("external_offer_id", "") or "").strip(),
                "vendor_code": str(item.get("vendor_code", product["art"]) or product["art"]).strip(),
                "external_name": str(item.get("external_name", product["name"]) or product["name"]).strip(),
            }

        self.conn.execute(
            "DELETE FROM product_barcodes WHERE product_id=? AND source='manual_mp'",
            (product_id,),
        )
        self.conn.execute(
            "DELETE FROM marketplace_product_links WHERE product_id=? AND raw_payload_json LIKE ?",
            (product_id, '%manual_mp_binding%'),
        )

        for code, item in normalized.items():
            barcode = item["barcode"]
            if barcode:
                self.conn.execute(
                    "INSERT OR IGNORE INTO product_barcodes(product_id, barcode, source, marketplace_code, is_primary) VALUES(?,?,?,?,1)",
                    (product_id, barcode, 'manual_mp', code),
                )
            self.conn.execute(
                "INSERT INTO marketplace_product_links(product_id, marketplace_code, external_product_id, external_sku, external_offer_id, vendor_code, external_name, raw_payload_json) VALUES(?,?,?,?,?,?,?,?)",
                (
                    product_id,
                    code,
                    '',
                    item["external_sku"],
                    item["external_offer_id"],
                    item["vendor_code"],
                    item["external_name"],
                    json.dumps({"seed": "manual_mp_binding"}, ensure_ascii=False),
                ),
            )

        bc_wb = normalized.get('wb', {}).get('barcode', '')
        bc_ozon = normalized.get('ozon', {}).get('barcode', '')
        self.conn.execute("UPDATE products SET bc_wb=?, bc_ozon=? WHERE id=?", (bc_wb, bc_ozon, product_id))
        self.conn.commit()

    # ── STOCK CALCULATIONS ──────────────────────────────────
    def get_stock(self, art):
        inc = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM arrivals WHERE art=?", (art,)).fetchone()[0]
        out = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM moves WHERE art=?", (art,)).fetchone()[0]
        return inc - out

    def get_stock_detail(self, art):
        inc   = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM arrivals WHERE art=?", (art,)).fetchone()[0]
        to_wb = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM moves WHERE art=? AND mp_type='WB'", (art,)).fetchone()[0]
        to_oz = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM moves WHERE art=? AND mp_type='OZON'", (art,)).fetchone()[0]
        to_ot = self.conn.execute("SELECT COALESCE(SUM(qty),0) FROM moves WHERE art=? AND (mp_type='' OR mp_type NOT IN ('WB','OZON'))", (art,)).fetchone()[0]
        return {"inc": inc, "wb": to_wb, "ozon": to_oz, "other": to_ot, "stock": inc - to_wb - to_oz - to_ot}

    def get_all_stock(self):
        rows = self.conn.execute("SELECT * FROM products ORDER BY art").fetchall()
        result = []
        for p in rows:
            d = self.get_stock_detail(p["art"])
            result.append({**dict(p), **d})
        return result

    # ── ARRIVALS ────────────────────────────────────────────
    def add_arrival(self, art, wh_name, qty, supplier="", note=""):
        self.conn.execute(
            "INSERT INTO arrivals(dt,art,wh_name,qty,supplier,note) VALUES(?,?,?,?,?,?)",
            (date.today().isoformat(), art, wh_name, qty, supplier, note))
        self.conn.commit()

    def get_arrivals(self, search="", limit=500):
        q = f"%{search}%"
        return self.conn.execute(
            "SELECT a.*, p.name as pname FROM arrivals a "
            "LEFT JOIN products p ON p.art=a.art "
            "WHERE a.art LIKE ? OR p.name LIKE ? OR a.wh_name LIKE ? "
            "ORDER BY a.id DESC LIMIT ?", (q, q, q, limit)).fetchall()

    def delete_arrival(self, aid):
        self.conn.execute("DELETE FROM arrivals WHERE id=?", (aid,))
        self.conn.commit()

    # ── MOVES ───────────────────────────────────────────────
    def add_move(self, art, from_wh, to_wh, qty, barcode="", mp_type="", note=""):
        self.conn.execute(
            "INSERT INTO moves(dt,art,from_wh,to_wh,qty,barcode,mp_type,note) VALUES(?,?,?,?,?,?,?,?)",
            (date.today().isoformat(), art, from_wh, to_wh, qty, barcode, mp_type, note))
        self.conn.commit()

    def get_moves(self, search="", mp_filter="", limit=500):
        q = f"%{search}%"
        mp_cond = ""
        params  = [q, q, q]
        if mp_filter == "WB":
            mp_cond = " AND m.mp_type='WB'"
        elif mp_filter == "OZON":
            mp_cond = " AND m.mp_type='OZON'"
        params.append(limit)
        return self.conn.execute(
            f"SELECT m.*, p.name as pname FROM moves m "
            f"LEFT JOIN products p ON p.art=m.art "
            f"WHERE (m.art LIKE ? OR p.name LIKE ? OR m.to_wh LIKE ?){mp_cond} "
            f"ORDER BY m.id DESC LIMIT ?", params).fetchall()

    def delete_move(self, mid):
        self.conn.execute("DELETE FROM moves WHERE id=?", (mid,))
        self.conn.commit()

    # ── LABEL TEMPLATES ─────────────────────────────────────
    def get_templates(self):
        return self.conn.execute("SELECT * FROM label_templates ORDER BY id").fetchall()

    def get_default_template(self):
        t = self.conn.execute("SELECT * FROM label_templates WHERE is_default=1").fetchone()
        if not t:
            t = self.conn.execute("SELECT * FROM label_templates LIMIT 1").fetchone()
        return t

    def save_template(self, tid, name, config_json, set_default=False):
        if tid:
            self.conn.execute("UPDATE label_templates SET name=?,config=? WHERE id=?",
                              (name, config_json, tid))
        else:
            self.conn.execute("INSERT INTO label_templates(name,config,is_default) VALUES(?,?,0)",
                              (name, config_json))
            tid = self.conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if set_default:
            self.conn.execute("UPDATE label_templates SET is_default=0")
            self.conn.execute("UPDATE label_templates SET is_default=1 WHERE id=?", (tid,))
        self.conn.commit()
        return tid

    def delete_template(self, tid):
        self.conn.execute("DELETE FROM label_templates WHERE id=?", (tid,))
        self.conn.commit()

    # ── XLSX EXPORT ─────────────────────────────────────────
    def export_xlsx(self, path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        def sheet(name, headers, rows):
            ws = wb.create_sheet(name)
            hfill = PatternFill("solid", start_color="1F2D3D")
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = Font(name="Arial", size=10)
                    if r % 2 == 0:
                        cell.fill = PatternFill("solid", start_color="F5F5F5")
            return ws

        del wb["Sheet"]

        sheet("Склады",
              ["Код", "Название", "Тип", "Примечание"],
              [(w["code"], w["name"], w["type"], w["note"]) for w in self.get_warehouses()])

        prods = self.get_products()
        sheet("Товары",
              ["Артикул", "Наименование", "Категория", "Ед.", "Мин.остаток",
               "Штрихкод WB", "Штрихкод OZON", "Поставщик", "Примечание"],
              [(p["art"], p["name"], p["cat"], p["unit"], p["min_stock"],
                p["bc_wb"], p["bc_ozon"], p["supplier"], p["note"]) for p in prods])

        sheet("Приход",
              ["Дата", "Артикул", "Наименование", "Склад", "Кол-во", "Поставщик", "Примечание"],
              [(a["dt"], a["art"], a["pname"] or "", a["wh_name"], a["qty"],
                a["supplier"], a["note"]) for a in self.get_arrivals(limit=9999)])

        sheet("Перемещения",
              ["Дата", "Артикул", "Наименование", "Откуда", "Куда", "Кол-во",
               "Штрихкод МП", "Тип МП", "Примечание"],
              [(m["dt"], m["art"], m["pname"] or "", m["from_wh"], m["to_wh"],
                m["qty"], m["barcode"], m["mp_type"], m["note"]) for m in self.get_moves(limit=9999)])

        stock = self.get_all_stock()
        sheet("Остатки",
              ["Артикул", "Наименование", "Кат.", "Ед.", "Приход",
               "→WB", "→OZON", "→др.", "Остаток", "Мин.", "Статус"],
              [(s["art"], s["name"], s["cat"], s["unit"], s["inc"],
                s["wb"], s["ozon"], s["other"], s["stock"], s["min_stock"],
                "НЕТ" if s["stock"] <= 0 else ("МАЛО" if s["min_stock"] > 0 and s["stock"] <= s["min_stock"] else "OK"))
               for s in stock])

        wb.save(path)

    # ── XLSX IMPORT ─────────────────────────────────────────
    def import_xlsx(self, path):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)

        def rows(sheet_name, skip=2):
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if not ws: return []
            return [[c.value for c in row] for row in ws.iter_rows(min_row=skip) if any(c.value for c in row)]

        def bc(v):
            if v is None: return ""
            if isinstance(v, float): return str(int(v))
            return str(v).strip()

        imported = {"products": 0, "arrivals": 0, "moves": 0, "warehouses": 0}

        for r in rows("📦 Товары"):
            if not r[0]: continue
            try:
                self.conn.execute(
                    "INSERT OR IGNORE INTO products(art,name,cat,unit,min_stock,bc_wb,bc_ozon,supplier,note) VALUES(?,?,?,?,?,?,?,?,?)",
                    (bc(r[0]), bc(r[1]), bc(r[2]), bc(r[3]) or "шт", int(r[4] or 0),
                     bc(r[5]), bc(r[6]), bc(r[7]), bc(r[9])))
                imported["products"] += 1
            except:
                pass

        for r in rows("🏭 Склады"):
            if not r[0]: continue
            try:
                self.conn.execute(
                    "INSERT OR IGNORE INTO warehouses(code,name,type,note) VALUES(?,?,?,?)",
                    (bc(r[0]), bc(r[1]), bc(r[2]) or "Собственный", bc(r[3])))
                imported["warehouses"] += 1
            except:
                pass

        for r in rows("➕ Приход"):
            if not r[1]: continue
            try:
                self.conn.execute(
                    "INSERT INTO arrivals(dt,art,wh_name,qty,supplier,note) VALUES(?,?,?,?,?,?)",
                    (bc(r[0]), bc(r[1]), bc(r[4]), int(r[5] or 0), bc(r[6]), bc(r[7])))
                imported["arrivals"] += 1
            except:
                pass

        for r in rows("🔄 Перемещения"):
            if not r[1]: continue
            try:
                self.conn.execute(
                    "INSERT INTO moves(dt,art,from_wh,to_wh,qty,barcode,mp_type,note) VALUES(?,?,?,?,?,?,?,?)",
                    (bc(r[0]), bc(r[1]), bc(r[4]), bc(r[5]), int(r[6] or 0), bc(r[7]), bc(r[8]), bc(r[9])))
                imported["moves"] += 1
            except:
                pass

        self.conn.commit()
        self._backfill_marketplace_foundation()
        self.conn.commit()
        return imported

    # ── SETTINGS ──────────────────────────────────────────
    def get_setting(self, key):
        row = self.conn.execute(
            "SELECT value FROM app_settings WHERE key=?", (key,)
        ).fetchone()
        return row["value"] if row else None

    def set_setting(self, key, value):
        self.conn.execute(
            "INSERT INTO app_settings(key,value) VALUES(?,?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value)
        )
        self.conn.commit()
